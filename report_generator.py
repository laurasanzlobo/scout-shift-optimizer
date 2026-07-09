# generador_informes.py
import io
from typing import List, Dict, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from jinja2 import Template
import weasyprint

# ─────────────────────────────────────────────────────────────────────────────
# Paleta de colores por rama scout
# ─────────────────────────────────────────────────────────────────────────────
COLORES_GRUPOS = {
    "Castores":  {"hex": "FF9800", "texto": "FFFFFF", "bg_ligero": "FFF3E0"},  
    "Lobatos":   {"hex": "FBC02D", "texto": "333333", "bg_ligero": "FFF9C4"},  
    "Ranger":    {"hex": "1E88E5", "texto": "FFFFFF", "bg_ligero": "E3F2FD"},  
    "Pioneros":  {"hex": "D32F2F", "texto": "FFFFFF", "bg_ligero": "FFEBEE"},  
    "Rutas":     {"hex": "43A047", "texto": "FFFFFF", "bg_ligero": "E8F5E9"},  
}

COLORES_TURNOS = {
    "Desayuno": "FFF9C4",
    "Comida":   "C8E6C9",
    "Cena":     "BBDEFB",
}

DIAS = [d for d in range(15, 31) if d not in {17, 18, 19}]
TURNOS = ["Desayuno", "Comida", "Cena"]
GRUPO_CASTORES_RUTAS = "CastoresRutas"


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _mapa_limpieza(limpieza: Optional[List[Dict]]) -> Dict:
    if not limpieza:
        return {}
    return {
        (item["Dia"], item["Turno"]): item["GrupoLimpieza"]
        for item in limpieza
    }


def _grupos_limpieza_para_mostrar(grupo: Optional[str]) -> List[str]:
  if not grupo:
    return []
  if grupo in {GRUPO_CASTORES_RUTAS}:
    return ["Castores", "Rutas"]
  return [grupo]


def _texto_limpieza_para_mostrar(grupo: Optional[str]) -> str:
  grupos = _grupos_limpieza_para_mostrar(grupo)
  return " / ".join(grupos) if grupos else "—"

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL (Se mantiene idéntico)
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel(asignaciones: List[Dict], limpieza: Optional[List[Dict]] = None) -> bytes:
    df = pd.DataFrame(asignaciones)
    limpieza_dict = _mapa_limpieza(limpieza)
    wb = Workbook()

    _hoja_por_dia(wb, df, limpieza_dict)
    _hoja_por_persona(wb, df)
    _hoja_datos(wb, df)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _estilo_cabecera(cell, color_hex: str, texto_hex: str = "FFFFFF"):
    cell.font = Font(bold=True, color=texto_hex, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=color_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()

def _hoja_por_dia(wb: Workbook, df: pd.DataFrame, limpieza_dict: Dict):
    ws = wb.create_sheet("Por Día")
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Turnos de Comedor — Campamento Verano 2026"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    cabeceras = ["Día", "Turno", "Responsables", "Grupos", "Limpieza"]
    for col, txt in enumerate(cabeceras, start=1):
        cell = ws.cell(row=3, column=col, value=txt)
        _estilo_cabecera(cell, "37474F")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18

    fila = 4
    for dia in DIAS:
        for turno in TURNOS:
            subset = df[(df["Dia"] == dia) & (df["Turno"] == turno)]
            if subset.empty:
                continue

            nombres = ", ".join(subset["Nombre"].tolist())
            grupos = ", ".join(subset["Grupo"].tolist())
            grupo_limpieza = limpieza_dict.get((dia, turno))
            grupo_limpieza_texto = _texto_limpieza_para_mostrar(grupo_limpieza)
            color_fondo = COLORES_TURNOS.get(turno, "FFFFFF")
            colores_limpieza = COLORES_GRUPOS.get(grupo_limpieza, {"hex": "ECEFF1", "texto": "333333"})

            valores = [dia, turno, nombres, grupos]
            for col, valor in enumerate(valores, start=1):
                cell = ws.cell(row=fila, column=col, value=valor)
                cell.fill = PatternFill("solid", start_color=color_fondo)
                cell.border = _thin_border()
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            celda_limpieza = ws.cell(row=fila, column=5, value=grupo_limpieza_texto)
            celda_limpieza.fill = PatternFill("solid", start_color=colores_limpieza["hex"])
            celda_limpieza.font = Font(name="Arial", size=9, bold=True, color=colores_limpieza["texto"])
            celda_limpieza.border = _thin_border()
            celda_limpieza.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.row_dimensions[fila].height = 30
            fila += 1

    ws.freeze_panes = "A4"

def _hoja_por_persona(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Por Persona")
    ws.merge_cells("A1:F1")
    ws["A1"] = "Carga de Turnos por Responsable"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    cabeceras = ["Nombre", "Grupo", "Nº Turnos", "Desayunos", "Comidas", "Cenas"]
    for col, txt in enumerate(cabeceras, start=1):
        cell = ws.cell(row=3, column=col, value=txt)
        _estilo_cabecera(cell, "37474F")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for col_letra in ["C", "D", "E", "F"]:
        ws.column_dimensions[col_letra].width = 12

    resumen = (df.groupby(["Nombre", "Grupo", "Turno"]).size().unstack(fill_value=0).reset_index())
    for t in TURNOS:
        if t not in resumen.columns:
            resumen[t] = 0
    resumen["Total"] = resumen[TURNOS].sum(axis=1)
    resumen = resumen.sort_values(["Grupo", "Nombre"])

    fila = 4
    for _, row in resumen.iterrows():
        grupo = row["Grupo"]
        colores = COLORES_GRUPOS.get(grupo, {"hex": "ECEFF1", "texto": "333333"})

        valores = [row["Nombre"], grupo, row["Total"], row.get("Desayuno", 0), row.get("Comida", 0), row.get("Cena", 0)]
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.fill = PatternFill("solid", start_color=colores["hex"])
            cell.font = Font(name="Arial", size=9, color=colores["texto"])
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        fila += 1

    total_row = fila
    ws.cell(row=total_row, column=1, value="TOTAL TURNOS").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})").font = Font(bold=True, name="Arial")
    for col in [4, 5, 6]:
        letra = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letra}4:{letra}{total_row-1})").font = Font(bold=True, name="Arial")

    ws.freeze_panes = "A4"

def _hoja_datos(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Datos")
    ws["A1"] = "Volcado de datos — generado automáticamente"
    ws["A1"].font = Font(italic=True, color="888888", name="Arial", size=9)

    cabeceras = ["Dia", "Turno", "Nombre", "Grupo"]
    for col, txt in enumerate(cabeceras, start=1):
        cell = ws.cell(row=2, column=col, value=txt)
        _estilo_cabecera(cell, "546E7A")

    for fila, (_, row) in enumerate(df.sort_values(["Dia", "Turno"]).iterrows(), start=3):
        for col, campo in enumerate(cabeceras, start=1):
            ws.cell(row=fila, column=col, value=row[campo])

    for col_letra in ["A", "B", "C", "D"]:
        ws.column_dimensions[col_letra].width = 18

# ─────────────────────────────────────────────────────────────────────────────
# PDF — Diseño "Presentación Diapositivas"
# ─────────────────────────────────────────────────────────────────────────────
_PLANTILLA_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
  /* Configuro el documento en A4 horizontal aprovechando bien los márgenes */
  @page {
    size: A4 landscape;
    margin: 1cm;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
    color: #333;
    font-size: 10pt;
  }

  /* ── COLORES DINÁMICOS POR RAMA ── */
  {% for grupo, col in colores.items() %}
  .bg-{{ grupo|lower|replace(' ','') }} { background: #{{ col.hex }}; color: #{{ col.texto }}; }
  {% endfor %}

  /* ── CONTENEDORES Y SALTOS DE PÁGINA ── */
  /* Separo las dos tablas en dos contenedores distintos. 
     Solo le aplico el salto de página al contenedor del comedor, 
     así evito que la primera página se quede en blanco. */
  .pagina {
    width: 100%;
    height: 100%;
    display: flex;
    flex-direction: column;
    justify-content: center;
  }
  .salto-pagina {
    page-break-before: always;
  }

  /* ── CABECERA ── */
  .header {
    text-align: center;
    margin-bottom: 0.6cm;
  }
  .header h1 {
    font-size: 22pt;
    font-weight: 800;
    color: #1C2331;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-bottom: 0.1cm;
  }
  .header p {
    font-size: 12pt;
    color: #7F8C8D;
  }
  
  .section-title {
    font-size: 14pt;
    font-weight: bold;
    color: #2C3E50;
    margin-bottom: 0.4cm;
    border-bottom: 2px solid #FF9800;
    display: inline-block;
    padding-bottom: 0.1cm;
  }

  /* ── TABLAS GLOBALES ── */
  table {
    width: 100%;
    border-collapse: collapse;
    /* Esto es vital: fuerza a que todas las columnas de los días midan exactamente lo mismo 
       y el texto no desborde ni deforme la celda. */
    table-layout: fixed; 
  }
  th, td {
    border: 1px solid #BDC3C7;
    text-align: center;
    vertical-align: middle;
    overflow: hidden;
  }
  th {
    background-color: #34495E;
    color: white;
    font-size: 10pt;
    padding: 0.25cm 0;
  }
  .th-turno {
    width: 8%; /* Hago la columna de los turnos un poco más estrecha que las de los días */
    background-color: #2C3E50;
    color: white;
    font-weight: bold;
  }

  /* ── TABLA LIMPIEZA (PÁGINA 1) ── */
  .td-limpieza {
    height: 1.8cm; 
    padding: 0.1cm;
  }
  /* He reducido drásticamente los márgenes y tamaños de fuente para que 
     quepan perfectamente Castores y Rutas juntos sin desbordar. */
  .badge-limpieza {
    display: inline-block;
    padding: 0.15cm 0.25cm;
    font-size: 8pt;
    font-weight: 700;
    border-radius: 0.15cm;
    margin: 0.05cm 0;
    white-space: nowrap; /* Evito que el nombre de un grupo se parta en dos líneas */
  }
  .vacio {
    color: #95A5A6;
    font-size: 8pt;
  }

  /* ── TABLA COMEDOR (PÁGINA 2) ── */
  .td-comedor {
    height: 4.5cm; /* Le doy más altura porque es una lista de personas */
    padding: 0.2cm;
    font-size: 10pt;
    text-align: center;
    vertical-align: middle;
    line-height: 1.4;
  }
</style>
</head>
<body>

<div class="pagina">
  <div class="header">
    <h1>Cuadrantes de Campamento</h1>
    <p>Julio 2026</p>
  </div>
  <div class="section-title">Equipos de Limpieza</div>
  
  <table>
    <thead>
      <tr>
        <th class="th-turno">Turno</th>
        {% for dia in dias %}
        <th>Día {{ dia }}</th>
        {% endfor %}
      </tr>
    </thead>
    <tbody>
      {% for turno in turnos %}
      <tr>
        <td class="th-turno">{{ turno }}</td>
        {% for dia in dias %}
        <td class="td-limpieza">
          {% set grupos = matriz_limpieza[turno][dia] %}
          {% if grupos %}
            {% for g in grupos %}
              <span class="badge-limpieza bg-{{ g|lower|replace(' ','') }}">{{ g }}</span>
            {% endfor %}
          {% else %}
            <span class="vacio">—</span>
          {% endif %}
        </td>
        {% endfor %}
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

<div class="pagina salto-pagina">
  <div class="header">
    <h1>Cuadrantes de Campamento</h1>
    <p>Julio 2026</p>
  </div>
  <div class="section-title">Servicio de Comedor</div>
  
  <table>
    <thead>
      <tr>
        <th class="th-turno">Turno</th>
        {% for dia in dias %}
        <th>Día {{ dia }}</th>
        {% endfor %}
      </tr>
    </thead>
    <tbody>
      {% for turno in turnos %}
      <tr>
        <td class="th-turno">{{ turno }}</td>
        {% for dia in dias %}
        <td class="td-comedor">
          {% set nombres = matriz_comedor[turno][dia] %}
          {% if nombres %}
            {{ nombres }}
          {% else %}
            <div style="text-align:center; color:#95A5A6; margin-top:0.5cm;">—</div>
          {% endif %}
        </td>
        {% endfor %}
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

</body>
</html>
"""

def generar_pdf(asignaciones: List[Dict], limpieza: Optional[List[Dict]] = None, anio: int = 2026) -> bytes:
    df = pd.DataFrame(asignaciones)
    limpieza_dict = _mapa_limpieza(limpieza)

    # Horas de referencia por turno, para mostrarlas junto al nombre en la cabecera de fila.
    horas_turno = {"Desayuno": "08:30", "Comida": "13:30", "Cena": "20:30"}

    # Aquí armo mi matriz de limpieza: turno -> dia -> grupo (o None si ese turno
    # no tiene servicio ese día, p.ej. la comida del 20 o la cena del 30).
    matriz_limpieza: Dict[str, Dict[int, List[str]]] = {
      turno: {dia: _grupos_limpieza_para_mostrar(limpieza_dict.get((dia, turno))) for dia in DIAS}
        for turno in TURNOS
    }

    # Y aquí armo mi matriz de comedor: turno -> dia -> string con los nombres
    # de los responsables ya unidos por comas, lista para pintar en la celda.
    matriz_comedor: Dict[str, Dict[int, Optional[str]]] = {turno: {} for turno in TURNOS}
    for turno in TURNOS:
        for dia in DIAS:
            subset = df[(df["Dia"] == dia) & (df["Turno"] == turno)] if not df.empty else df
            if subset.empty:
                matriz_comedor[turno][dia] = None
            else:
                matriz_comedor[turno][dia] = ", ".join(subset["Nombre"].tolist())

    # Inyecto todas las matrices y diccionarios, incluyendo la paleta de colores
    html_renderizado = Template(_PLANTILLA_HTML).render(
        dias=DIAS,
        turnos=TURNOS,
        horas_turno=horas_turno,
        matriz_limpieza=matriz_limpieza,
        matriz_comedor=matriz_comedor,
        colores=COLORES_GRUPOS  # <-- Esta es la variable que me faltaba
    )

    pdf_bytes = weasyprint.HTML(string=html_renderizado).write_pdf()
    return pdf_bytes